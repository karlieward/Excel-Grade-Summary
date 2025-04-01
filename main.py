# Karlie Ward, Rachel Pinkney, Sabrina Wong, 
# Spencer Bigelow, Mason Zarges, Gavin Smith

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Creates a student class to hold all of the information
class Student:
    
    def __init__(self, className, grade, last, first, studentID):
        self.id = studentID
        self.first = first
        self.last = last
        self.grade = grade
        self.className = className

# Open the excel sheet to extract information
wbOriginal = openpyxl.load_workbook("Poorly_Organized_Data_2.xlsx")
wsOriginal = wbOriginal.active

# List of student objects
lstStudents = []

# Class Dictionary
dictionaryClass = {}

# Sets active row as the second row to avoid the column headers
activeRow = 2

# Runs through the original sheet as long as there are values in the column
while True:
    activeCell = wsOriginal.cell(row=activeRow, column=1).value  # Get value of the current cell in column A
    if activeCell is None:  # Stop if cell is empty
        break

    class_name = activeCell
    studentInfo = wsOriginal.cell(row=activeRow, column=2).value
    last, first, studentID = wsOriginal.cell(row=activeRow, column=2).value.split('_')
    studentGrade = wsOriginal.cell(row=activeRow, column=3).value

    oStudent = Student(class_name, studentGrade, last, first, studentID) # Create new student 
    lstStudents.append(oStudent)  # Append student to the list

    if class_name in dictionaryClass:
        dictionaryClass[class_name].append(oStudent)
    else:
        dictionaryClass[class_name] = []
        dictionaryClass[class_name].append(oStudent)

    activeRow += 1

# Set up the beginning columns for each new sheet
myWorkbook = Workbook()

for key in dictionaryClass:
    newWS = myWorkbook.create_sheet(key)
    newWS["A1"] = "Last Name"
    newWS["B1"] = "First Name"
    newWS["C1"] = "Student ID"
    newWS["D1"] = "Grade"
    newWS["F1"] = "Summary Statistics"
    newWS["G1"] = "Value"

    count = 2
    for student in dictionaryClass[key]:
        newWS.cell(row=count, column=1).value = student.last
        newWS.cell(row=count, column=2).value = student.first
        newWS.cell(row=count, column=3).value = student.id
        newWS.cell(row=count, column=4).value = student.grade
        
        count += 1

# Get cell range for headers
cell_range = ['A1', 'B1', 'C1', 'D1']

# Create a bold font to apply to headers
bold_font = Font(bold=True)

# Iterate through each worksheet accessing by index
for i in range(5):
    worksheet = myWorkbook.worksheets[i]
    
    for cell in cell_range:
        header_cell = worksheet[cell]
        header_cell.font = bold_font  # Apply bold font
        
        # Adjust olumn width from text length + 5 units
        if header_cell.value:
            column_letter = cell[0]
            column_width = len(header_cell.value) + 5
            worksheet.column_dimensions[column_letter].width = column_width

# Delete default "Sheet"
deleteSheet = myWorkbook["Sheet"]
myWorkbook.remove(deleteSheet)

    newWS.auto_filter.ref = "A:D"

    count -= 1

    newWS["F2"] = "Highest Grade"
    newWS["G2"] = f"=MAX(D2:D{count})"

    newWS["F3"] = "Lowest Grade"
    newWS["G3"] = f"=MIN(D2:D{count})"

    newWS["F4"] = "Mean Grade"
    newWS["G4"] = f"=AVERAGE(D2:D{count})"
    
    newWS["F5"] = "Median Grade"
    newWS["G5"] = f"=MEDIAN(D2:D{count})"
    
    newWS["F6"] = "Number of Students"
    newWS["G6"] = f"=COUNT(D2:D{count})"


myWorkbook.save(filename="formatted_grades.xlsx")
myWorkbook.close()