# this is gavin's part

import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font


class Student :
    def __init__(self, student_class, last_name, first_name, ID, grade) :
        self.class_name = student_class
        self.first_name = first_name
        self.last_name = last_name
        self.stud_ID = ID
        self.grade = grade


class Class :
    def __init__(self, class_name) :
        self.class_name = class_name
        self.students


gradebook = Workbook()

gradebook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

currSheet = gradebook.active

lstClasses = []
lstClassObj = []
# currSheet.insert_rows(1,1)
# currSheet.insert_cols(3, 3)
# currSheet["C1"] = "Last Name"
# currSheet["D1"] = "First Name"
# currSheet["E1"] = "Student ID"

gradebook.create_sheet("Algebra")
# currSheet["A5"] = "Test"

currSheet.auto_filter.ref = "A:F"

# calculate the last row 
lastRow = max((cell.row for cell in currSheet["B"] if cell.value), default=1)


for row_index, (value, ) in enumerate(currSheet.iter_rows(min_row=2, max_row=lastRow, min_col=1, max_col=1, values_only=True), start=2) :
    sClass = str(currSheet.cell(row=row_index, column=1).value)

    if sClass not in lstClasses :
        oClass = Class(sClass)
        lstClasses.append(sClass)
        oClass = Class(sClass)
        lstClassObj.append(oClass)


for row_index, (value, ) in enumerate(currSheet.iter_rows(min_row = 2, max_row = lastRow, min_col = 2, max_col = 2, values_only = True), start=3) :

    parts = value.split("_")
    print(parts)
        
    if len(parts) == 3 :

        # oStud = Student()

        # lName = currSheet.cell(row=row_index, column = 3, value = parts[0])
        # fName = currSheet.cell(row=row_index, column = 4, value = parts[1])
        # sID = currSheet.cell(row=row_index, column = 5, value = parts[2])
        
        lName = parts[0]
        fName = parts[1]
        sID = parts[2]
        sClass = currSheet.cell(row = row_index, column = 1).value
        iGrade = currSheet.cell(row = row_index, column = 3).value

    oStud = Student(sClass, lName, fName, sID, iGrade)







gradebook.save(filename = "Working_Book.xlsx")
gradebook.close()